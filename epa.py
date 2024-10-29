#Python 3.12.2 

#Imports
from pathlib import Path
import os, logging, openpyxl
from openpyxl.styles import PatternFill
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
import openpyxl.workbook
from utility import load_file_as_base64
import configparser
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from datetime import datetime

#Create a log.file
logging.basicConfig(
        filename='mi_log.log',              # Nombre del archivo donde se guardarán los logs
        level=logging.DEBUG,                # Nivel de logging (DEBUG, INFO, WARNING, ERROR, CRITICAL)
        format='%(asctime)s - %(levelname)s - %(message)s',  # Formato del mensaje
        datefmt='%Y-%m-%d %H:%M:%S'         # Formato de la fecha y hora
    )

def verificarDir(path):
    try: 
        dir = Path(path)
        logging.debug(f'La carpeta objetivo {dir} existe')
        return dir
    except:
        logging.error(f'La carpeta NO objetivo {dir} existe')
        return False
    
def load_config():
    try:
        config = configparser.ConfigParser()
        config.read('client.ini')
        
        excel_path = config.get('Settings', 'excel_path')
        input_folder = config.get('Settings', 'input_folder')
        
        return excel_path, input_folder
    except Exception as e:
        logging.error(f'Error al cargar la configuración: {e}')
        return None, None



def verificarExcel():
    try:
        excel_path, input_folder = load_config()
        excel = openpyxl.load_workbook(excel_path)
        logging.debug(f'Se pudo encontrar el excel')
        
        
        return excel ,input_folder, excel_path
        
    except FileNotFoundError:
        logging.error('No se pudo encontrar el archivo Excel. Verifica la ruta.')
    except openpyxl.utils.exceptions.InvalidFileException as e:
        logging.error(f'El archivo {excel_path} no es un formato de Excel válido: {e}')
    except Exception as e:
        logging.error(f'Error al abrir el excel: {e}')
    
    return False, False, False


def createAzureSession():
    
    try:
        logging.debug(f'Iniciando sesion de Azure')
        # Leer variables del config.ini
        config = configparser.ConfigParser()
        config.read('client.ini')

        # Asignar las variables de Azure
        endpoint = config.get('DocumentAI', 'endpoint')
        api_key = config.get('DocumentAI', 'api_key')

        # Inicializar el Document Intelligence client
        credential = AzureKeyCredential(api_key)
        client = DocumentIntelligenceClient(endpoint=endpoint, credential=credential)

        logging.debug(f'Sesion de Azure creada')
        return client
    except configparser.NoSectionError as e:
        logging.error(f'Error en la configuración: {e}')
    except configparser.NoOptionError as e:
        logging.error(f'Opción faltante en la configuración: {e}')
    except Exception as e:
        logging.error(f'Error al crear la sesión de Azure: {e}', exc_info=True)
    
    return False

    
def proccesDoc(path, client, model_id):
    try:
        # Cargar el archivo y convertirlo a base64
        file_base64 = load_file_as_base64(path)
        
        # Iniciar el análisis del documento con locale en español
        poller = client.begin_analyze_document(
            model_id,
            {"base64Source": file_base64},
            locale="es-ES",  # Cambiado a español (España)
        )
        
        # Puedes agregar lógica para esperar el resultado aquí, si es necesario
        return poller
    except Exception as e:
        logging.error(f'Error processing document: {e}', exc_info=True)
        return False
    

def iterarArchivos(dir, excel, excel_path):

    client = createAzureSession()
    amarillo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    sheet = excel.active

    for file in os.listdir(dir):
        path = os.path.join(dir, file)
        if os.path.isfile(path):
            try: 
                logging.debug(f'Archivo {file} fue correctamente encontrado.')
                result = proccesDoc(path, client, 'EPA6')
                logging.debug(result.result().documents)
                if result:
                    logging.debug(f'Archivo {file} fue correctamente leido por Azure DocumentAI.')
                    row = 0
                    for doc in result.result().documents:
                        fields = doc['fields']
                        logging.debug(fields)
                        ctg = fields.get('CTG')
                        pesoNeto = fields.get('PesoNeto')
                        provinciaProcedencia = fields.get('ProvinciaProcedencia')
                        fechaDescarga = fields.get('FechaDescarga')
                        observaciones = fields.get('Observaciones')
                        localidadProcedencia = fields.get('LocalidadProcedencia')

                        #logging.debug(f'ctg: {ctg['content']}, pesoneto {pesoNeto['content']}, prov: {provinciaDestino['content']}, fecha: {fechaDescarga['content']}, obs: {observaciones['content']}, local: {localidadDestino['content']}')

                        if ctg:
                            for fila in sheet.iter_rows(min_col=5, max_col=5):  
                                    for celda in fila:
                                        if str(celda.value) in str(ctg['content']):
                                            row = celda.row
                                            celda.fill = amarillo
                                            logging.debug(f'El valor {celda.value} fue pintado.')
                                            
                                            
                                            if pesoNeto:
                                                try:
                                                    if str(sheet[f'F{row}'].value) in str(pesoNeto['content']):
                                                            sheet[f'F{row}'].fill = amarillo
                                                            logging.debug(f'El valor {sheet[f'F{row}'].value} fue pintado.')
                                                    else:
                                                        logging.debug(f'El valor del pdf {pesoNeto['content']} difiere del excel {sheet[f'F{row}'].value}.')  
                                                except:
                                                    continue
                                            
                                            if provinciaProcedencia:
                                                try:    
                                                    if str(sheet[f'C{row}'].value) in str(provinciaProcedencia['content']):
                                                            sheet[f'C{row}'].fill = amarillo
                                                            logging.debug(f'El valor {sheet[f'C{row}'].value} fue pintado.')
                                                    else:
                                                        logging.debug(f'El valor del pdf {provinciaProcedencia['content']} difiere del excel {sheet[f'C{row}'].value}.')  
                                                except:
                                                    continue
                                            
                                            if fechaDescarga:
                                                try:
                                                    # Convierte la fecha de 'dd/mm/yyyy' a un objeto datetime
                                                    fecha_descarga_date = datetime.strptime(fechaDescarga['content'], '%d/%m/%Y').date()
                                                    

                                                    # Obtén el valor de la celda y conviértelo a un objeto datetime
                                                    valor_celda = sheet[f'D{row}'].value
                                                    fecha = valor_celda.strftime('%Y-%m-%d') 

                                                    if str(fecha_descarga_date) in str(fecha):
                                                            sheet[f'D{row}'].fill = amarillo
                                                            logging.debug(f'El valor {sheet[f'D{row}'].value} fue pintado.')
                                                    else:
                                                        logging.debug(f'El valor del pdf {fecha_descarga_date} difiere del excel {fecha}.')              
                                                except:
                                                    continue
                                            
                                            if observaciones:
                                                try:
                                                    if str(sheet[f'A{row}'].value) in str(observaciones['content']):
                                                            sheet[f'A{row}'].fill = amarillo
                                                            logging.debug(f'El valor {sheet[f'A{row}'].value} fue pintado.')
                                                    else:
                                                        logging.debug(f'El valor del pdf {observaciones['content']} difiere del excel {sheet[f'A{row}'].value}.')  
                                                except:
                                                    continue                    
                                            
                                            if localidadProcedencia:
                                                try:
                                                    if str(sheet[f'B{row}'].value) in str(localidadProcedencia['content']):
                                                            sheet[f'B{row}'].fill = amarillo
                                                            logging.debug(f'El valor {sheet[f'B{row}'].value} fue pintado.')
                                                    else:
                                                        logging.debug(f'El valor del pdf {localidadProcedencia['content']} difiere del excel {sheet[f'B{row}'].value}.')  
                                                except:
                                                    continue


                                        excel.save(excel_path)
            except:
                continue
        else:
            logging.debug(f'Archivo {file} fue NO correctamente leido por Azure DocumentAI.')
        


def main():
    
    excel, folder, excel_path = verificarExcel()

    if excel:
        iterarArchivos(verificarDir(folder), excel, excel_path)
    else:
        print('Fail')
main()
    

